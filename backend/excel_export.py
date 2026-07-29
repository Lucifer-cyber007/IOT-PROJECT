"""Builds the .xlsx export of extracted bills."""

from __future__ import annotations

import io
import re
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# (key, column heading, width). Order defines the sheet layout.
COLUMNS: list[tuple[str, str, int]] = [
    ("source_file", "Source File", 26),
    ("name", "Name", 26),
    ("rr_number", "RR Number", 16),
    ("account_number", "Account Number", 18),
    ("address", "Address", 42),
    ("units_consumed", "Units Consumed", 16),
    ("amount_to_pay", "Amount to be Paid", 18),
    ("tariff", "Tariff", 14),
    ("bill_date", "Bill Date", 14),
]

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# Columns Excel should treat as numbers rather than text.
NUMERIC_KEYS = {"units_consumed", "amount_to_pay"}


def _as_number(value: str) -> float | int | str:
    """Convert a numeric-looking string so Excel can sum/sort it."""
    candidate = value.replace(",", "").strip()
    if not re.fullmatch(r"-?\d+(\.\d+)?", candidate):
        return value
    number = float(candidate)
    return int(number) if number.is_integer() and "." not in candidate else number


def build_workbook(rows: list[dict], filename: str) -> StreamingResponse:
    """Render rows into an .xlsx download response."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bills"

    sheet.append([heading for _, heading, _ in COLUMNS])
    for index, (_, _, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = width
        cell = sheet[f"{letter}1"]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    sheet.row_dimensions[1].height = 22

    for row in rows:
        values: list[object] = []
        for key, _, _ in COLUMNS:
            raw = row.get(key)
            text = "" if raw is None else str(raw).strip()
            values.append(_as_number(text) if key in NUMERIC_KEYS and text else text)
        sheet.append(values)

    # Right-align the numeric columns and give the address room to wrap.
    for key, _, _ in COLUMNS:
        column_index = [c[0] for c in COLUMNS].index(key) + 1
        letter = get_column_letter(column_index)
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet[f"{letter}{row_index}"]
            if key in NUMERIC_KEYS:
                cell.alignment = Alignment(horizontal="right")
                if key == "amount_to_pay" and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
            elif key == "address":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Freeze the header and add filters so a long batch stays navigable.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{sheet.max_row}"

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # RFC 5987 encoding so non-ASCII filenames survive the header.
    ascii_name = re.sub(r"[^A-Za-z0-9._-]", "_", filename) or "bills.xlsx"
    disposition = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"

    return StreamingResponse(
        stream,
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": disposition,
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
